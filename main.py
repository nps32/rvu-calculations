# import statements
import openpyxl
import subprocess
from openpyxl import Workbook
from dataclasses import dataclass

# data class obj for Procedure
@dataclass
class Procedure:
    doctor: str
    proc_code: int
    proc_name: str
    rvu_value: float


doctors = list()
procedureCategories = list()
all_procedures = list()

# Create dict to hold pro
procedureCodes = dict()

# function to read in the list of process codes stored in file 'proc-codes.xlsx'
def readProcedureCodes():
    # load the file into a Workbook obj
    processWorkbook = openpyxl.load_workbook('proc-codes.xlsx')
    processSheet = processWorkbook.active

    # Add code-identifier key-value pairs to dict
    for column in processSheet.iter_cols():
        procedureCategories.append(column[0].value)
        for cell in column:
            if (cell.value is not None) and (cell.value != column[0].value):
                procedureCodes[cell.value] = column[0].value

    procedureCodes[93248] = "DummyVal"

    # for key, value in procedureCodes.items():
    # print(f"Key: {key}, Value: {value}")
    # print("There are " + len(proc_dict) + " process codes)

    return procedureCodes

def addProcCodestoMasterSheet(masterWorkbook, procedureCodes):
    masterSheet = masterWorkbook.active
    totalRows = masterSheet.max_row + 1
    print("There are", totalRows, "rows total in masterSheet")

    for i in range(2, totalRows):
        row = masterSheet[i]
        procCode = row[1].value

        if procCode is not None:
            if procCode in procedureCodes:
                procCategory = procedureCodes.get(procCode)
                row[6].value = procCategory
            else:
                print("Error finding procedure category for procedure code", procCode, "on row", i)

    masterWorkbook.save("mastersheet.xlsx")

# function to append all individual sheet data to master copy
def populateMasterSheet(masterWorkbook, sheetnames):
    masterSheet = masterWorkbook.active

    for filename in sheetnames:
        currentWorkbook = openpyxl.load_workbook(filename)
        currentSheet = currentWorkbook.active
        print(filename + " has", currentSheet.max_row, "total rows of data")
        doctors.append(currentSheet["A2"].value)

        sheet_last_row = currentSheet.max_row - 1

        for i in range(2, sheet_last_row):
            row = currentSheet[i]

            # Creating an instance of the Procedure class
            procedure_instance = Procedure(row[0].value, row[1].value, row[2].value, row[3].value)
            all_procedures.append(procedure_instance)

            rowContents = list()

            for cell in row:
                rowContents.append(cell.value)
                # print(cell.value, "  ", end='')
            # print()

            masterSheet.append(rowContents)

    masterWorkbook.save("mastersheet.xlsx")



def makeSecondarySheet(masterWorkbook):
    rvuSheet = masterWorkbook.create_sheet(title="RVU Sums")

    firstrow = list()
    firstrow.append("Doctors")
    firstrow += procedureCategories

    rvuSheet.append(firstrow)

    number_doctors = len(doctors)
    print("There are", number_doctors, "doctors total")

    for i in range(2, (number_doctors + 1)):
        row = list()

        # Add doctor
        current_doctor = doctors[i-2]
        row.append(doctors[i-2])

        for j in range (0, len(procedureCategories)):
            category_sum = 0
            for k in range (0, len(all_procedures)):
                if all_procedures[k].doctor == current_doctor and procedureCodes[all_procedures[k].proc_code] == procedureCategories[i]:
                    category_sum += int(all_procedures[k].rvu_value)
                row.append(str(category_sum))

        rvuSheet.append(row)

    masterWorkbook.save("mastersheet.xlsx")


# function to create the master sheet
def makeMasterSheet(sheetnames):
    master_workbook = Workbook()
    master_sheet = master_workbook.active

    master_sheet["A1"] = "PHYSICIAN NAME"
    master_sheet["B1"] = "PROC_CODE"
    master_sheet["C1"] = "PROC_NAME"
    master_sheet["D1"] = "Sum of CHARGES"
    master_sheet["E1"] = "Sum of PROC_QTY"
    master_sheet["F1"] = "Sum of RVU VALUE"
    master_sheet["G1"] = "PROC_CATEGORY"

    # Append all values from individual spreadsheets
    populateMasterSheet(master_workbook, sheetnames)
    addProcCodestoMasterSheet(master_workbook, readProcedureCodes())
    makeSecondarySheet(master_workbook)


def run_bash_script(script_path):
    try:
        # Run the bash script using subprocess.run()
        result = subprocess.run(['bash', script_path], capture_output=True, text=True, check=True)

        # If the script runs successfully, the output will be stored in result.stdout
        return result.stdout
    except subprocess.CalledProcessError as e:
        # If the script encounters an error and returns a non-zero exit code, you can handle it here
        print(f"Error executing the script: {e}")
        return None


def split_input_string(input_string):
    # Split the input string
    parts = input_string.strip().split('\n')

    # Remove any empty strings resulting from consecutive newlines
    parts = [part.strip() for part in parts if part.strip()]

    return parts


def getFileNames():
    output = run_bash_script('list_xlsx_files.sh')
    sheetnames = split_input_string(output)

    if "proc-codes.xlsx" in sheetnames:
        sheetnames.remove("proc-codes.xlsx")

    if "mastersheet.xlsx" in sheetnames:
        sheetnames.remove("mastersheet.xlsx")

    print("filenames retrieved by bash script: ", end='')
    for file_name in sheetnames:
        print(file_name, end='  ')
    print()

    return sheetnames


sheetnames = getFileNames()
makeMasterSheet(sheetnames)

for procedure in procedureCategories:
    print(procedure)
